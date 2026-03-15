"""app/utils/export_excel.py
Export container packing results to a formatted Excel report.

Sheets:
  1. Overview        — summary metrics + per-container fill table + bar chart
  2. Container Details — full breakdown per container (all items sorted by Y position)
  3. Packing Layout  — colour-coded cell-grid (pallets + boxes on same row, rec sub-row)
  4. Recommendations — what to add to fill free space
"""

from pathlib import Path
from typing import Dict, List, Any, Optional
import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import SeriesLabel
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


# ═══════════════════════════════════════════════════════════════════════════
# Colour palette
# ═══════════════════════════════════════════════════════════════════════════

_HEADER_BG    = "1A3A5C"   # dark navy
_SUBHDR_BG    = "2E6DA4"   # mid blue
_ACCENT_BG    = "D6E4F0"   # pale blue
_ALT_ROW      = "EBF5FB"   # very light blue
_WHITE        = "FFFFFF"
_LIGHT_GRAY   = "F5F5F5"

_BLOCK_PAL: Dict[str, str] = {
    "115x115": "AED6F1",   # sky blue
    "115x108": "FAD7A0",   # peach
    "115x77":  "A9DFBF",   # mint green
    "77x77":   "D7BDE2",   # lavender
}
_NP_BOX_COLOR  = "FFF176"  # vivid yellow      — NP box zones
_EMPTY_COLOR   = "E8E8E8"  # light grey        — unused space
_REC_PAL_COLOR = "43A047"  # strong green      — recommended pallets
_REC_NP_COLOR  = "00838F"  # strong teal       — recommended NP boxes
_REC_ROW_BG    = "F1F8E9"  # pale green tint   — rec sub-row background
_OPTIMAL_BG    = "1B5E20"  # dark green        — "already optimal" header
_HADD_BG       = "E65100"  # deep orange       — "additions possible" header

_TL_GOOD = "27AE60"
_TL_OK   = "E67E22"

_FOOTPRINT_TO_PALLET_TYPE: Dict[str, str] = {
    "115x115": "A2",
    "115x108": "A1",
    "115x77":  "C2",
    "77x77":   "D2",
}


# ═══════════════════════════════════════════════════════════════════════════
# Low-level style helpers
# ═══════════════════════════════════════════════════════════════════════════

def _fill(hex_color: str) -> "PatternFill":
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10, italic=False) -> "Font":
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def _align(h="left", v="center", wrap=False) -> "Alignment":
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _thin_border() -> "Border":
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_top_border() -> "Border":
    t = Side(style="medium", color="43A047")
    s = Side(style="thin", color="D0D0D0")
    return Border(top=t, left=s, right=s, bottom=s)

def _block_color(key: str) -> str:
    for prefix, color in _BLOCK_PAL.items():
        if key.startswith(prefix):
            return color
    return "CCCCCC"

def _has_recommendations(rec: dict) -> bool:
    return (rec.get("total_pallets_to_add", 0) > 0
            or rec.get("total_np_boxes_to_add", 0) > 0)


def _vol_fill_pct(container: dict, W: int, Hdoor: int, L: int) -> float:
    pallet_vol = sum(
        r["length_cm"] * W * r["height_cm"]
        for r in container.get("rows", [])
    )
    box_zone_vol = sum(
        p["length_cm"] * p["width_cm"] * p["height_cm"] * p["quantity"]
        for z in container.get("box_zones", [])
        for p in z.get("placed", [])
    )
    container_vol = float(L * W * Hdoor)
    if container_vol <= 0:
        return 0.0
    return round(100.0 * (pallet_vol + box_zone_vol) / container_vol, 1)


def _set_cell(ws, row: int, col: int, value=None, *,
              bold=False, fg="000000", bg=None, size=10, italic=False,
              h="left", v="center", wrap=False,
              border=False, num_fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = _font(bold=bold, color=fg, size=size, italic=italic)
    cell.alignment = _align(h=h, v=v, wrap=wrap)
    if bg:
        cell.fill = _fill(bg)
    if border:
        cell.border = _thin_border()
    if num_fmt:
        cell.number_format = num_fmt
    return cell


def _merge_write(ws, r1, c1, r2, c2, value="", *,
                 bold=False, fg="000000", bg=None, size=10, italic=False,
                 h="left", v="center"):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=value)
    cell.font      = _font(bold=bold, color=fg, size=size, italic=italic)
    cell.alignment = _align(h=h, v=v)
    if bg:
        cell.fill = _fill(bg)
    return cell


def _section_title(ws, row: int, c1: int, c2: int, title: str, bg=_SUBHDR_BG):
    _merge_write(ws, row, c1, row, c2, title,
                 bold=True, fg=_WHITE, bg=bg, size=11, h="left")
    ws.row_dimensions[row].height = 22


def _col_header_row(ws, row: int, headers: List[str], col_start: int = 1,
                    widths: Optional[List[float]] = None):
    for i, h in enumerate(headers):
        c = col_start + i
        _set_cell(ws, row, c, h, bold=True, bg=_ACCENT_BG, fg=_HEADER_BG,
                  h="center", border=True)
        if widths and i < len(widths):
            ws.column_dimensions[get_column_letter(c)].width = widths[i]
    ws.row_dimensions[row].height = 18


# ═══════════════════════════════════════════════════════════════════════════
# Sheet 1 — Overview
# ═══════════════════════════════════════════════════════════════════════════

def _write_overview(ws, containers, recs, np_boxes, unplaced, config):
    ws.sheet_view.showGridLines = False

    L     = config.get("CONTAINER_LENGTH_CM", 1203)
    W     = config.get("CONTAINER_WIDTH_CM", 235)
    Hdoor = config.get("CONTAINER_DOOR_HEIGHT_CM", 250)

    total_pallets  = sum(c["loaded_value"]  for c in containers)
    total_weight   = sum(c["loaded_weight"] for c in containers)
    total_np_boxes = sum(
        sum(p["quantity"] for z in c.get("box_zones", []) for p in z["placed"])
        for c in containers
    )
    total_rec_pal  = sum(r.get("total_pallets_to_add", 0)  for r in recs)
    total_rec_np   = sum(r.get("total_np_boxes_to_add", 0) for r in recs)
    total_unplaced = sum(e["remaining_qty"] for e in (unplaced or []))

    avg_fill = (
        sum(_vol_fill_pct(c, W, Hdoor, L) for c in containers) / len(containers)
        if containers else 0.0
    )

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 18

    _merge_write(ws, 1, 1, 1, 8,
                 "🚢  CONTAINER PACKING REPORT",
                 bold=True, fg=_WHITE, bg=_HEADER_BG, size=16, h="center")
    ws.row_dimensions[1].height = 38

    run_time = datetime.datetime.now().strftime("%d %b %Y  %H:%M")
    _merge_write(ws, 2, 1, 2, 8,
                 f"Generated: {run_time}   |   Objective: {config.get('RECOMMEND_OBJECTIVE', '—')}",
                 fg="666666", size=9, h="center", bg=_LIGHT_GRAY)
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 8

    _section_title(ws, 4, 1, 8, "  KEY METRICS")

    metrics = [
        ("Containers used",             len(containers)),
        ("Total pallets loaded",         total_pallets),
        ("Total NP boxes loaded",        total_np_boxes),
        ("Total weight loaded (kg)",     f"{total_weight:,.0f}"),
        ("Avg container volume fill",    f"{avg_fill:.1f} %"),
        ("Recommended extra pallets",    total_rec_pal),
        ("Recommended extra NP boxes",   total_rec_np),
        ("Unplaced NP boxes",            total_unplaced),
    ]

    for i, (label, value) in enumerate(metrics):
        r = 5 + i
        bg = _ALT_ROW if i % 2 == 0 else _WHITE
        _set_cell(ws, r, 1, label, bold=True, bg=bg, border=True)
        _set_cell(ws, r, 2, value, bg=bg, h="right", border=True)
        _merge_write(ws, r, 3, r, 8, bg=bg)
        ws.row_dimensions[r].height = 17

    ws.row_dimensions[13].height = 10

    _section_title(ws, 14, 1, 8, "  CONTAINER SUMMARY")

    hdrs = ["Container", "Pallets used (cm)", "Total (cm)", "Vol Fill %",
            "Pallets", "Weight (kg)", "NP Boxes", "Rec Pallets"]
    _col_header_row(ws, 15, hdrs, col_start=1,
                    widths=[12, 14, 12, 10, 10, 14, 12, 14])

    rec_by_idx = {r["container_index"]: r for r in recs}
    chart_rows_start = 16

    for i, c in enumerate(containers):
        r      = 16 + i
        idx    = c["container_index"]
        used   = c["used_length_cm"]
        pals   = c["loaded_value"]
        wt     = c["loaded_weight"]
        zones  = c.get("box_zones", [])
        boxes  = sum(p["quantity"] for z in zones for p in z["placed"])
        rec    = rec_by_idx.get(idx, {})
        r_pal  = rec.get("total_pallets_to_add", 0)
        fill_p = _vol_fill_pct(c, W, Hdoor, L)

        bg = _ALT_ROW if i % 2 == 0 else _WHITE
        data = [idx, used, L, fill_p, pals, round(wt, 0), boxes, r_pal]
        for j, val in enumerate(data):
            col  = 1 + j
            cell = _set_cell(ws, r, col, val, bg=bg, h="center", border=True)
            if j == 3:
                well_filled = fill_p >= 85
                cell.number_format = '0.0"%"'
                cell.fill = _fill(_TL_GOOD if well_filled else _TL_OK)
                cell.font = _font(bold=True, color=_WHITE, size=10)
            elif j in (1, 2, 5):
                cell.number_format = '#,##0'
        ws.row_dimensions[r].height = 17

    chart_rows_end = 16 + len(containers) - 1

    try:
        chart = BarChart()
        chart.type         = "col"
        chart.grouping     = "clustered"
        chart.title        = "Container Volume Fill %"
        chart.y_axis.title = "Vol Fill %"
        chart.x_axis.title = "Container"
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 100
        chart.height = 12
        chart.width  = 18

        data_ref = Reference(ws, min_col=4, max_col=4,
                             min_row=chart_rows_start, max_row=chart_rows_end)
        cats_ref = Reference(ws, min_col=1, max_col=1,
                             min_row=chart_rows_start, max_row=chart_rows_end)
        chart.add_data(data_ref, from_rows=False, titles_from_data=False)
        chart.set_categories(cats_ref)
        chart.series[0].title = SeriesLabel(v="Vol Fill %")
        chart.series[0].graphicalProperties.solidFill = "2E6DA4"

        anchor_row = chart_rows_end + 3
        ws.add_chart(chart, f"A{anchor_row}")
    except Exception:
        pass


# ═══════════════════════════════════════════════════════════════════════════
# Sheet 2 — Container Details
# ═══════════════════════════════════════════════════════════════════════════

def _write_details(ws, containers, config):
    """
    Unified table per container — all items (pallet rows + NP box placements)
    sorted by Y-start position so the length accounting is crystal clear.
    """
    ws.sheet_view.showGridLines = False

    col_widths = [12, 22, 14, 12, 12, 14, 10]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    L     = config.get("CONTAINER_LENGTH_CM", 1203)
    W     = config.get("CONTAINER_WIDTH_CM", 235)
    Hdoor = config.get("CONTAINER_DOOR_HEIGHT_CM", 250)

    HEADERS = ["Type", "Product / Block", "Y Start (cm)",
               "Length (cm)", "Height (cm)", "Weight (kg)", "Count"]

    row = 1
    for c in containers:
        idx    = c["container_index"]
        pals   = c["loaded_value"]
        wt     = c["loaded_weight"]
        vol_p  = _vol_fill_pct(c, W, Hdoor, L)
        zones  = c.get("box_zones", [])
        n_boxes = sum(p["quantity"] for z in zones for p in z["placed"])

        pallet_used = c["used_length_cm"]
        box_len     = sum(z["length_cm"] for z in zones)
        empty_len   = max(0, L - pallet_used - box_len)

        # ── Container header ──────────────────────────────────────────────
        _merge_write(ws, row, 1, row, 7,
                     f"  CONTAINER {idx}",
                     bold=True, fg=_WHITE, bg=_HEADER_BG, size=12)
        ws.row_dimensions[row].height = 24
        row += 1

        # Space accounting bar (one info row)
        acct = (f"  Pallets: {pallet_used} cm   |   NP Boxes: {box_len} cm   |   "
                f"Empty: {empty_len} cm   |   Total: {L} cm   "
                f"||   Pallets loaded: {pals}   "
                f"NP boxes loaded: {n_boxes}   Weight: {wt:,.0f} kg   "
                f"Vol fill: {vol_p}%")
        _merge_write(ws, row, 1, row, 7, acct,
                     fg="333333", size=9, bg=_ACCENT_BG)
        ws.row_dimensions[row].height = 16
        row += 1

        # ── Column headers ────────────────────────────────────────────────
        _col_header_row(ws, row, HEADERS, col_start=1)
        row += 1

        # ── Build unified item list sorted by Y position ──────────────────
        items = []

        for rrow in c.get("rows", []):
            bk = rrow["block_type"]
            fp = bk.split("|")[0] if "|" in bk else bk
            pt = _FOOTPRINT_TO_PALLET_TYPE.get(fp, "—")
            items.append({
                "y": rrow["y_start_cm"],
                "type_label": f"Pallet  [{pt}]",
                "product": f"{bk}",
                "length": rrow["length_cm"],
                "height": rrow["height_cm"],
                "weight": round(rrow["weight_kg"], 1),
                "count": rrow["pallet_count"],
                "is_pallet": True,
                "block_key": bk,
            })

        for zone in zones:
            for placed in zone["placed"]:
                items.append({
                    "y": zone["y_start_cm"],
                    "type_label": "NP Box",
                    "product": placed["label"],
                    "length": placed["length_cm"],
                    "height": placed["height_cm"],
                    "weight": round(placed["weight_kg_total"], 1),
                    "count": placed["quantity"],
                    "is_pallet": False,
                    "block_key": "",
                })

        items.sort(key=lambda x: x["y"])

        for ri, item in enumerate(items):
            bg = _ALT_ROW if ri % 2 == 0 else _WHITE
            color_type = _block_color(item["block_key"]) if item["is_pallet"] else _NP_BOX_COLOR
            vals = [
                item["type_label"],
                item["product"],
                item["y"],
                item["length"],
                item["height"],
                item["weight"],
                item["count"],
            ]
            for ci, v in enumerate(vals):
                cell_bg = color_type if ci in (0, 1) else bg
                _set_cell(ws, row, ci + 1, v,
                          bg=cell_bg, border=True,
                          h="center" if ci >= 2 else "left",
                          wrap=(ci == 1))
            ws.row_dimensions[row].height = 16
            row += 1

        # Empty space row
        if empty_len > 0:
            _set_cell(ws, row, 1, "Empty space", bg=_EMPTY_COLOR, border=True, italic=True, fg="888888")
            _set_cell(ws, row, 2, "— no items —", bg=_EMPTY_COLOR, border=True, italic=True, fg="888888")
            _set_cell(ws, row, 3, pallet_used + box_len, bg=_EMPTY_COLOR, border=True, h="center", fg="888888")
            _set_cell(ws, row, 4, empty_len, bg=_EMPTY_COLOR, border=True, h="center", fg="888888")
            for ci in range(5, 8):
                _set_cell(ws, row, ci, "—", bg=_EMPTY_COLOR, border=True, h="center", fg="888888")
            ws.row_dimensions[row].height = 16
            row += 1

        ws.row_dimensions[row].height = 10  # spacer
        row += 1


# ═══════════════════════════════════════════════════════════════════════════
# Sheet 3 — Packing Layout  (colour-coded cell grid)
# ═══════════════════════════════════════════════════════════════════════════

_CM_PER_COL        = 12
_LAYOUT_LABEL_COLS = 3   # cols 1-3: Container | Fill% | MaxH


def _cm_to_col(cm: int) -> int:
    return _LAYOUT_LABEL_COLS + 1 + (cm // _CM_PER_COL)

def _layout_col_count(L_cm: int) -> int:
    return (L_cm // _CM_PER_COL) + 1


def _color_layout_range(ws, row: int, y_start_cm: int, length_cm: int,
                        color: str, label: str = "", border_top: bool = False):
    c_start = _cm_to_col(y_start_cm)
    c_end   = max(_cm_to_col(y_start_cm + length_cm - 1), c_start)
    fill    = _fill(color)
    border  = _thick_top_border() if border_top else None

    for c in range(c_start, c_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        if border:
            cell.border = border

    if label and (c_end - c_start) >= 1:
        lbl_cell = ws.cell(row=row, column=c_start, value=label)
        lbl_cell.font      = _font(bold=True, color="222222", size=7)
        lbl_cell.alignment = _align(h="left", v="center")


def _scale_row_height(max_h_cm: int) -> int:
    """Map block height in cm to a row pixel height for the layout grid."""
    # 270 cm container maps to 56px; floor at 22px
    return max(22, min(56, int(max_h_cm * 56 / 270)))


def _write_layout(ws, containers, recs, config):
    ws.sheet_view.showGridLines = False

    L     = config.get("CONTAINER_LENGTH_CM", 1203)
    W     = config.get("CONTAINER_WIDTH_CM", 235)
    Hdoor = config.get("CONTAINER_DOOR_HEIGHT_CM", 250)
    n_layout_cols = _layout_col_count(L)
    total_cols    = _LAYOUT_LABEL_COLS + n_layout_cols + 1

    # ── Column widths ─────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 12   # Container #
    ws.column_dimensions["B"].width = 9    # Fill %
    ws.column_dimensions["C"].width = 9    # Max height

    for c in range(_LAYOUT_LABEL_COLS + 1, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 1.8

    # ── Title ─────────────────────────────────────────────────────────────
    _merge_write(ws, 1, 1, 1, total_cols,
                 "  PACKING LAYOUT   "
                 "each column ≈ 12 cm  •  pallets + NP boxes shown on same row  •  "
                 "green/teal sub-row = recommended additions",
                 bold=True, fg=_WHITE, bg=_HEADER_BG, size=11, h="left")
    ws.row_dimensions[1].height = 26

    # ── Legend ────────────────────────────────────────────────────────────
    legend_items = [
        ("115×115 pallet", _BLOCK_PAL["115x115"]),
        ("115×108 pallet", _BLOCK_PAL["115x108"]),
        ("115×77  pallet", _BLOCK_PAL["115x77"]),
        ("77×77   pallet", _BLOCK_PAL["77x77"]),
        ("NP boxes",       _NP_BOX_COLOR),
        ("Empty space",    _EMPTY_COLOR),
        ("Rec. pallets",   _REC_PAL_COLOR),
        ("Rec. NP boxes",  _REC_NP_COLOR),
    ]
    ws.row_dimensions[2].height = 4
    leg_row = 3
    ws.row_dimensions[leg_row].height = 16
    col = 1
    for label, color in legend_items:
        swatch = ws.cell(row=leg_row, column=col, value="  " + label)
        swatch.fill      = _fill(color)
        swatch.font      = _font(size=8, color="222222" if color not in (_REC_PAL_COLOR, _REC_NP_COLOR) else "FFFFFF")
        swatch.alignment = _align(h="left", v="center")
        swatch.border    = _thin_border()
        try:
            ws.merge_cells(start_row=leg_row, start_column=col,
                           end_row=leg_row, end_column=col + 1)
        except Exception:
            pass
        col += 3

    ws.row_dimensions[4].height = 4

    # ── Ruler ─────────────────────────────────────────────────────────────
    ruler_row = 5
    ws.row_dimensions[ruler_row].height = 13
    _set_cell(ws, ruler_row, 1, "Container",  bold=True, size=8, h="center", bg=_ACCENT_BG)
    _set_cell(ws, ruler_row, 2, "Vol Fill %", bold=True, size=8, h="center", bg=_ACCENT_BG)
    _set_cell(ws, ruler_row, 3, "Max H (cm)", bold=True, size=8, h="center", bg=_ACCENT_BG)

    for tick_cm in range(0, L + 1, 100):
        c = _cm_to_col(tick_cm)
        if c <= total_cols:
            cell = ws.cell(row=ruler_row, column=c, value=str(tick_cm))
            cell.font      = _font(size=7, color="555555", bold=True)
            cell.alignment = _align(h="center", v="center")
            cell.fill      = _fill(_ACCENT_BG)

    # ── Container rows ────────────────────────────────────────────────────
    rec_by_idx = {r["container_index"]: r for r in recs}
    cur_row    = 6

    for c in containers:
        idx       = c["container_index"]
        fill_p    = _vol_fill_pct(c, W, Hdoor, L)
        rec       = rec_by_idx.get(idx, {})
        tl        = _TL_GOOD if fill_p >= 85 else _TL_OK
        rows_data = c.get("rows", [])
        zones     = c.get("box_zones", [])

        tail_pls = rec.get("tail_placements", [])
        atop_pls = rec.get("atop_placements", [])
        has_recs = bool(tail_pls or atop_pls)

        # Max block height drives row height (visual height cue)
        all_heights = [r["height_cm"] for r in rows_data] or [0]
        max_h = max(all_heights)
        main_row_h = _scale_row_height(max_h)

        main_row = cur_row
        rec_row  = (cur_row + 1) if has_recs else None
        n_rows   = 1 + (1 if has_recs else 0)

        # ── Label columns (merged across main + rec row) ───────────────────
        _merge_write(ws, main_row, 1,
                     main_row + n_rows - 1, 1,
                     f"Cont. {idx}",
                     bold=True, fg=_WHITE, bg=_SUBHDR_BG, size=9,
                     h="center", v="center")

        _merge_write(ws, main_row, 2,
                     main_row + n_rows - 1, 2,
                     fill_p,
                     bold=True, fg=_WHITE,
                     bg=tl, size=9, h="center", v="center")
        ws.cell(row=main_row, column=2).number_format = '0.0"%"'

        # Col 3: height label on main row only (do NOT merge into rec_row)
        _set_cell(ws, main_row, 3, f"↕ {max_h} cm",
                  size=7, fg="444444", bg=_LIGHT_GRAY, h="center", v="center")

        if rec_row:
            _set_cell(ws, rec_row, 3, "▲ REC",
                      size=7, bold=True, fg=_WHITE, bg=_REC_PAL_COLOR, h="center")

        # ── Row heights ────────────────────────────────────────────────────
        ws.row_dimensions[main_row].height = main_row_h
        if rec_row:
            ws.row_dimensions[rec_row].height = 16

        # ── Background: shade entire container length as empty ─────────────
        for layout_c in range(_LAYOUT_LABEL_COLS + 1, _cm_to_col(L) + 1):
            ws.cell(row=main_row, column=layout_c).fill = _fill(_EMPTY_COLOR)
            if rec_row:
                ws.cell(row=rec_row, column=layout_c).fill = _fill(_REC_ROW_BG)

        # ── Pallet blocks (main row) ───────────────────────────────────────
        for rrow in rows_data:
            bk    = rrow["block_type"]
            fp    = bk.split("|")[0] if "|" in bk else bk
            color = _block_color(bk)
            label = f"{fp}({rrow['pallet_count']}p) h:{rrow['height_cm']}cm"
            _color_layout_range(ws, main_row,
                                rrow["y_start_cm"], rrow["length_cm"],
                                color, label)

        # ── NP box zones (main row — same row as pallets) ──────────────────
        for zone in zones:
            total_boxes = sum(p["quantity"] for p in zone["placed"])
            _color_layout_range(ws, main_row,
                                zone["y_start_cm"], zone["length_cm"],
                                _NP_BOX_COLOR,
                                f"NP({total_boxes}b)")

        # ── Recommendations (rec sub-row with thick green top border) ──────
        if has_recs and rec_row:
            for p in tail_pls + atop_pls:
                y0    = p.get("y_start_cm", 0)
                lp    = p.get("length_cm", 0)
                is_np = p.get("type") == "np_box"
                color = _REC_NP_COLOR if is_np else _REC_PAL_COLOR
                units = p.get("units_per_placement", p.get("pallets_per_block", 0))
                lbl   = f"+{units}{'b' if is_np else 'p'}"
                _color_layout_range(ws, rec_row, y0, lp, color, lbl,
                                    border_top=True)

        # Spacer
        spacer = main_row + n_rows
        ws.row_dimensions[spacer].height = 5
        cur_row = spacer + 1


# ═══════════════════════════════════════════════════════════════════════════
# Sheet 4 — Recommendations
# ═══════════════════════════════════════════════════════════════════════════

def _write_recommendations(ws, recs):
    ws.sheet_view.showGridLines = False

    col_widths = [10, 20, 38, 14, 14, 10, 14, 14, 14]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    _merge_write(ws, 1, 1, 1, 9,
                 "  FILL RECOMMENDATIONS — additional items to order to fill free container space",
                 bold=True, fg=_WHITE, bg=_HEADER_BG, size=12)
    ws.row_dimensions[1].height = 28

    _merge_write(ws, 2, 1, 2, 9,
                 "  Tail Zone = unused length after last pallet row   |   "
                 "Atop Zone = headroom above existing pallet rows",
                 fg="444444", size=9, bg=_LIGHT_GRAY)
    ws.row_dimensions[2].height = 16

    row = 4

    for rec in recs:
        idx      = rec["container_index"]
        before   = rec.get("leftover_before_cm", 0)
        after    = rec.get("leftover_after_cm", 0)
        n_pal    = rec.get("total_pallets_to_add", 0)
        n_np     = rec.get("total_np_boxes_to_add", 0)
        ts       = rec.get("tail_summary",  {})
        as_      = rec.get("atop_summary",  {})

        has_any  = _has_recommendations(rec)
        before_m = round(before / 100, 2)

        # ── Container header ──────────────────────────────────────────────
        if not has_any:
            # Optimal container — no recommendations
            _merge_write(ws, row, 1, row, 9,
                         f"  ✓  CONTAINER {idx}  —  CURRENTLY OPTIMAL",
                         bold=True, fg=_WHITE, bg=_OPTIMAL_BG, size=11)
            ws.row_dimensions[row].height = 22
            row += 1

            empty_msg = (
                f"  {before} cm ({before_m} m) of tail space remains empty. "
                f"No additional items from the current order can fit given packing constraints."
                if before > 0 else
                "  Tail: fully packed — no unused space."
            )
            _merge_write(ws, row, 1, row, 9, empty_msg,
                         fg="2E7D32", size=9, bg="F1F8E9", italic=(before == 0))
            ws.row_dimensions[row].height = 18
            row += 1
            ws.row_dimensions[row].height = 6
            row += 1
            continue

        # Container has recommendations
        _merge_write(ws, row, 1, row, 9,
                     f"  ▲  CONTAINER {idx}  —  +{n_pal} PALLETS  +{n_np} NP BOXES CAN BE ADDED",
                     bold=True, fg=_WHITE, bg=_HADD_BG, size=11)
        ws.row_dimensions[row].height = 22
        row += 1

        space_msg = (
            f"  Tail before recommendations: {before} cm ({before_m} m)   →   "
            f"Remaining after: {after} cm"
            if before > 0 else "  Tail: already full before recommendations."
        )
        _merge_write(ws, row, 1, row, 9, space_msg,
                     fg="BF360C", size=9, bg="FBE9E7")
        ws.row_dimensions[row].height = 16
        row += 1

        # Column headers
        _col_header_row(ws, row,
                        ["Zone", "Type", "Product / Block", "Length (cm)", "Height (cm)",
                         "Count", "Units/Row", "Total Units", "Est. FOB"],
                        col_start=1)
        row += 1

        # ── PALLETS subsection ────────────────────────────────────────────
        pal_tail   = ts.get("pallet_blocks", [])
        pal_atop   = as_.get("pallet_blocks", [])
        all_pal    = pal_tail + pal_atop

        _merge_write(ws, row, 1, row, 9,
                     "  PALLETS TO ADD",
                     bold=True, fg=_WHITE, bg=_REC_PAL_COLOR, size=10)
        ws.row_dimensions[row].height = 18
        row += 1

        if all_pal:
            for ri, b in enumerate(pal_tail):
                _write_pallet_row(ws, row, b, "TAIL", ri)
                row += 1
            for ri, b in enumerate(pal_atop):
                _write_pallet_row(ws, row, b, "ATOP", ri + len(pal_tail))
                row += 1
        else:
            _merge_write(ws, row, 1, row, 9,
                         "   0 pallets recommended — no pallet blocks from current order fit the available space",
                         fg="777777", italic=True, bg=_LIGHT_GRAY, size=9)
            ws.row_dimensions[row].height = 16
            row += 1

        # ── NP BOXES subsection ───────────────────────────────────────────
        np_tail  = ts.get("np_box_rows", [])
        np_atop  = as_.get("np_box_rows", [])
        all_np   = np_tail + np_atop

        _merge_write(ws, row, 1, row, 9,
                     "  NP BOXES TO ADD",
                     bold=True, fg=_WHITE, bg=_REC_NP_COLOR, size=10)
        ws.row_dimensions[row].height = 18
        row += 1

        if all_np:
            for ri, b in enumerate(np_tail):
                _write_np_row(ws, row, b, "TAIL", ri)
                row += 1
            for ri, b in enumerate(np_atop):
                _write_np_row(ws, row, b, "ATOP", ri + len(np_tail))
                row += 1
        else:
            _merge_write(ws, row, 1, row, 9,
                         "   0 NP boxes recommended — no NP box types from current order fit the available space",
                         fg="777777", italic=True, bg=_LIGHT_GRAY, size=9)
            ws.row_dimensions[row].height = 16
            row += 1

        # ── Grand total row ───────────────────────────────────────────────
        ws.row_dimensions[row].height = 4
        row += 1

        _merge_write(ws, row, 1, row, 6,
                     f"   ▲ TOTAL ADDITIONS FOR CONTAINER {idx}",
                     bold=True, fg=_WHITE, bg=_SUBHDR_BG, size=10)
        _merge_write(ws, row, 7, row, 8,
                     f"+{n_pal} pallets",
                     bold=True, fg=_WHITE, bg=_REC_PAL_COLOR, size=10, h="center")
        _set_cell(ws, row, 9,
                  f"+{n_np} NP boxes",
                  bold=True, fg=_WHITE, bg=_REC_NP_COLOR, size=10, h="center")
        ws.row_dimensions[row].height = 20
        row += 1

        ws.row_dimensions[row].height = 10   # spacer
        row += 1


def _write_pallet_row(ws, row: int, b: dict, zone_label: str, ri: int):
    bg = _ALT_ROW if ri % 2 == 0 else _WHITE
    fob_s = f"{b['est_value_fob']:,.2f}" if b.get("est_value_fob") else "—"
    block_label = f"{b['block_type_key']}  —  {b['pallets_per_block']} pallets/block"
    vals = [
        zone_label, "Pallet block", block_label,
        b["length_cm"], b["height_cm"],
        b["count"], b["pallets_per_block"], b["total_pallets"], fob_s,
    ]
    for ci, v in enumerate(vals):
        color = _block_color(b["block_type_key"]) if ci == 2 else \
                _REC_PAL_COLOR if ci == 0 else bg
        _set_cell(ws, row, ci + 1, v, bg=color, border=True,
                  h="center" if ci > 2 else "left", wrap=(ci == 2),
                  fg="FFFFFF" if ci == 0 else "000000")
    ws.row_dimensions[row].height = 18


def _write_np_row(ws, row: int, b: dict, zone_label: str, ri: int):
    bg = _ALT_ROW if ri % 2 == 0 else _WHITE
    dim = b.get("dim", "")
    product_label = f"{b['label']}  [{dim}]" if dim else b["label"]
    vals = [
        zone_label, f"NP box  ({b['n_across']}× across width)", product_label,
        b["length_cm"], b["height_cm"],
        b["count"], b["n_across"], b["total_boxes"], "—",
    ]
    for ci, v in enumerate(vals):
        color = _NP_BOX_COLOR if ci == 2 else \
                _REC_NP_COLOR if ci == 0 else bg
        _set_cell(ws, row, ci + 1, v, bg=color, border=True,
                  h="center" if ci > 2 else "left", wrap=(ci == 2),
                  fg="FFFFFF" if ci == 0 else "000000")
    ws.row_dimensions[row].height = 28


# ═══════════════════════════════════════════════════════════════════════════
# Public API
# ═══════════════════════════════════════════════════════════════════════════

def export_excel_report(
    containers: List[Dict[str, Any]],
    recs: List[Dict[str, Any]],
    np_boxes: Optional[List[Dict[str, Any]]],
    unplaced: Optional[List[Dict[str, Any]]],
    out_dir: "Path",
    config: Dict[str, Any],
) -> Optional["Path"]:
    if not _HAS_OPENPYXL:
        print("[export_excel] openpyxl not installed — skipping Excel report")
        return None

    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("Overview")
    ws2 = wb.create_sheet("Container Details")
    ws3 = wb.create_sheet("Packing Layout")
    ws4 = wb.create_sheet("Recommendations")

    ws1.sheet_properties.tabColor = "1A3A5C"
    ws2.sheet_properties.tabColor = "2E6DA4"
    ws3.sheet_properties.tabColor = "27AE60"
    ws4.sheet_properties.tabColor = "E67E22"

    _write_overview(ws1, containers, recs, np_boxes or [], unplaced or [], config)
    _write_details(ws2, containers, config)
    _write_layout(ws3, containers, recs, config)
    _write_recommendations(ws4, recs)

    out_path = out_dir / "report.xlsx"
    wb.save(str(out_path))
    return out_path
